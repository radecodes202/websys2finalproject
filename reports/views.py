from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, F
from django.views.generic import TemplateView, DetailView
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
from accounts.mixins import RoleRequiredMixin
from accounts.models import User
from product.models import Sale, SaleItem, StockMovement, Product
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from xhtml2pdf import pisa
from django.template.loader import render_to_string
from io import BytesIO


class SalesReportView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/sales_report.html'
    allowed_roles = ['admin', 'manager']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        period = request.GET.get('period', 'day')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        qs = Sale.objects.select_related('cashier').prefetch_related('items')

        today = timezone.now().date()
        if period == 'day' and not date_from and not date_to:
            qs = qs.filter(date__date=today)
        elif period == 'week' and not date_from and not date_to:
            start_of_week = today - timezone.timedelta(days=today.weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            qs = qs.filter(date__date__gte=start_of_week, date__date__lte=end_of_week)
        elif period == 'month' and not date_from and not date_to:
            qs = qs.filter(date__year=today.year, date__month=today.month)
        elif date_from and date_to:
            qs = qs.filter(date__date__gte=date_from, date__date__lte=date_to)

        context['sales'] = qs.order_by('-date')
        context['period'] = period
        context['date_from'] = date_from
        context['date_to'] = date_to

        # Totals
        totals = qs.aggregate(
            total_sales=Sum('total'),
            total_subtotal=Sum('subtotal'),
            total_tax=Sum('tax'),
            total_discount=Sum('discount'),
            sale_count=Count('id'),
        )
        context.update(totals)

        # Profit = sum of (unit_price - cost_price) * quantity for completed sales
        from django.db.models import ExpressionWrapper, DecimalField
        profit_expr = ExpressionWrapper(
            F('items__unit_price') - F('items__product__cost_price'),
            output_field=DecimalField()
        )
        completed_qs = qs.filter(status=Sale.STATUS_COMPLETED)
        profit_agg = completed_qs.annotate(
            item_profit=profit_expr
        ).aggregate(total_profit=Sum('item_profit'))
        context['total_profit'] = profit_agg['total_profit'] or 0

        return context


class SalesHistoryView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/sales_history.html'
    allowed_roles = ['admin', 'manager', 'cashier']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        base_qs = Sale.objects.select_related('cashier').prefetch_related('items', 'payments')
        if request.user.role == 'cashier':
            base_qs = base_qs.filter(cashier=request.user)

        search = request.GET.get('search', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        cashier_id = request.GET.get('cashier', '').strip()
        payment_method = request.GET.get('payment_method', '').strip()
        status = request.GET.get('status', '').strip()

        if search:
            base_qs = base_qs.filter(
                Q(pk__icontains=search) |
                Q(customer__name__icontains=search)
            )
        if date_from:
            base_qs = base_qs.filter(date__date__gte=date_from)
        if date_to:
            base_qs = base_qs.filter(date__date__lte=date_to)
        if cashier_id and request.user.role in ('admin', 'manager'):
            base_qs = base_qs.filter(cashier_id=cashier_id)
        if payment_method:
            base_qs = base_qs.filter(payment_method=payment_method)
        if status:
            base_qs = base_qs.filter(status=status)

        page_size = request.GET.get('page_size', 20)
        paginator = Paginator(base_qs.order_by('-date'), page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        payment_methods = Sale.objects.values_list('payment_method', flat=True).distinct().order_by('payment_method')
        statuses = [choice[0] for choice in Sale.STATUS_CHOICES]
        users = User.objects.filter(is_active=True).order_by('username')

        context.update({
            'page_obj': page_obj,
            'payment_methods': payment_methods,
            'statuses': statuses,
            'users': users,
            'request_get': request.GET,
        })
        return context


class SaleDetailView(RoleRequiredMixin, DetailView):
    model = Sale
    template_name = 'reports/sale_detail.html'
    context_object_name = 'sale'
    allowed_roles = ['admin', 'manager', 'cashier']

    def get_queryset(self):
        qs = Sale.objects.select_related('cashier').prefetch_related('items', 'payments')
        if self.request.user.role == 'cashier':
            return qs.filter(cashier=self.request.user)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['print_url'] = reverse_lazy('product:receipt_print', kwargs={'pk': self.object.pk})
        context['export_pdf_url'] = reverse_lazy('reports:sale-pdf', kwargs={'pk': self.object.pk})
        context['export_excel_url'] = reverse_lazy('reports:sale-excel', kwargs={'pk': self.object.pk})
        return context


class InventoryValuationView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/inventory_valuation.html'
    allowed_roles = ['admin', 'manager']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = Product.objects.select_related('category').filter(is_active=True)
        rows = []
        total_value = 0
        for product in products:
            value = product.cost_price * product.quantity_in_stock
            rows.append({
                'product': product,
                'quantity': product.quantity_in_stock,
                'cost_price': product.cost_price,
                'value': value,
            })
            total_value += value
        context['rows'] = rows
        context['total_value'] = total_value
        context['export_excel_url'] = reverse_lazy('reports:inventory-excel')
        context['export_pdf_url'] = reverse_lazy('reports:inventory-pdf')
        return context


class StockMovementView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/stock_movement.html'
    allowed_roles = ['admin', 'manager']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        qs = StockMovement.objects.select_related('product')
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        product_id = request.GET.get('product', '').strip()
        movement_type = request.GET.get('type', '').strip()

        if date_from:
            qs = qs.filter(timestamp__date__gte=date_from)
        if date_to:
            qs = qs.filter(timestamp__date__lte=date_to)
        if product_id:
            qs = qs.filter(product_id=product_id)
        if movement_type:
            qs = qs.filter(type=movement_type)

        page_size = request.GET.get('page_size', 50)
        paginator = Paginator(qs.order_by('-timestamp'), page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context.update({
            'page_obj': page_obj,
            'date_from': date_from,
            'date_to': date_to,
            'products': Product.objects.filter(is_active=True).order_by('name'),
            'movement_types': StockMovement.MOVEMENT_TYPE_CHOICES,
        })
        return context


class SupplierPaymentOutstandingView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/supplier_payment_outstanding.html'
    allowed_roles = ['admin', 'manager']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Placeholder: aggregates outstanding supplier payments.
        # Integrates with SupplierPayment model if present; otherwise shows 0.
        try:
            from supplier.models import SupplierPayment
            payments = SupplierPayment.objects.filter(is_settled=False).select_related('supplier')
            rows = []
            total_outstanding = 0
            for payment in payments:
                rows.append({
                    'supplier': payment.supplier,
                    'amount': payment.amount,
                    'due_date': payment.due_date,
                    'reference': payment.reference_number,
                })
                total_outstanding += payment.amount
            context['rows'] = rows
            context['total_outstanding'] = total_outstanding
        except Exception:
            context['rows'] = []
            context['total_outstanding'] = 0
        return context


class ProfitLossSummaryView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/profit_loss_summary.html'
    allowed_roles = ['admin', 'manager']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        period = request.GET.get('period', 'month')
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()

        qs = Sale.objects.filter(status=Sale.STATUS_COMPLETED)

        today = timezone.now().date()
        if period == 'day' and not date_from and not date_to:
            qs = qs.filter(date__date=today)
        elif period == 'week' and not date_from and not date_to:
            start_of_week = today - timezone.timedelta(days=today.weekday())
            end_of_week = start_of_week + timezone.timedelta(days=6)
            qs = qs.filter(date__date__gte=start_of_week, date__date__lte=end_of_week)
        elif period == 'month' and not date_from and not date_to:
            qs = qs.filter(date__year=today.year, date__month=today.month)
        elif date_from and date_to:
            qs = qs.filter(date__date__gte=date_from, date__date__lte=date_to)

        revenue = qs.aggregate(total=Sum('total'))['total'] or 0

        # COGS from purchase stock movements in period
        cogs_qs = StockMovement.objects.filter(type=StockMovement.MOVEMENT_TYPE_PURCHASE)
        if period == 'day' and not date_from and not date_to:
            cogs_qs = cogs_qs.filter(timestamp__date=today)
        elif period == 'week' and not date_from and not date_to:
            cogs_qs = cogs_qs.filter(timestamp__date__gte=start_of_week, timestamp__date__lte=end_of_week)
        elif period == 'month' and not date_from and not date_to:
            cogs_qs = cogs_qs.filter(timestamp__year=today.year, timestamp__month=today.month)
        elif date_from and date_to:
            cogs_qs = cogs_qs.filter(timestamp__date__gte=date_from, timestamp__date__lte=date_to)

        from django.db.models import ExpressionWrapper, DecimalField
        cogs_expr = ExpressionWrapper(F('quantity_change') * F('product__cost_price'), output_field=DecimalField())
        cogs = cogs_qs.annotate(line_cost=cogs_expr).aggregate(total=Sum('line_cost'))['total'] or 0

        # Placeholder for expenses: supplier payments in period
        expenses = 0
        try:
            from supplier.models import SupplierPayment
            expense_qs = SupplierPayment.objects.all()
            if period == 'day' and not date_from and not date_to:
                expense_qs = expense_qs.filter(payment_date__date=today)
            elif period == 'week' and not date_from and not date_to:
                expense_qs = expense_qs.filter(payment_date__date__gte=start_of_week, payment_date__date__lte=end_of_week)
            elif period == 'month' and not date_from and not date_to:
                expense_qs = expense_qs.filter(payment_date__year=today.year, payment_date__month=today.month)
            elif date_from and date_to:
                expense_qs = expense_qs.filter(payment_date__date__gte=date_from, payment_date__date__lte=date_to)
            expenses = expense_qs.aggregate(total=Sum('amount'))['total'] or 0
        except Exception:
            pass

        net_profit = revenue - cogs - expenses

        context.update({
            'period': period,
            'date_from': date_from,
            'date_to': date_to,
            'revenue': revenue,
            'cogs': cogs,
            'expenses': expenses,
            'net_profit': net_profit,
        })
        return context


class SalePDFView(RoleRequiredMixin, DetailView):
    model = Sale
    allowed_roles = ['admin', 'manager', 'cashier']

    def get(self, request, *args, **kwargs):
        sale = self.get_object()
        template_path = 'reports/sale_pdf.html'
        html = render_to_string(template_path, {'sale': sale})
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sale_{sale.pk}.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response


class SaleExcelView(RoleRequiredMixin, DetailView):
    model = Sale
    allowed_roles = ['admin', 'manager', 'cashier']

    def get(self, request, *args, **kwargs):
        sale = self.get_object()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Sale {sale.pk}'
        ws.append(['Sale ID', 'Date', 'Cashier', 'Customer', 'Payment Method', 'Subtotal', 'Tax', 'Discount', 'Total', 'Status'])
        ws.append([
            sale.pk, sale.date.strftime('%Y-%m-%d %H:%M'), sale.cashier.username if sale.cashier else '',
            sale.customer.name if sale.customer else '', sale.payment_method,
            sale.subtotal, sale.tax, sale.discount, sale.total, sale.status
        ])
        ws.append([])
        ws.append(['Items'])
        ws.append(['Product', 'Quantity', 'Unit Price', 'Subtotal'])
        for item in sale.items.all():
            ws.append([item.product.name, item.quantity, item.unit_price, item.subtotal])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sale_{sale.pk}.xlsx"'
        wb.save(response)
        return response


class InventoryPDFView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/inventory_pdf.html'
    allowed_roles = ['admin', 'manager']

    def get(self, request, *args, **kwargs):
        products = Product.objects.select_related('category').filter(is_active=True)
        rows = []
        total_value = 0
        for product in products:
            value = product.cost_price * product.quantity_in_stock
            rows.append({'product': product, 'quantity': product.quantity_in_stock, 'cost_price': product.cost_price, 'value': value})
            total_value += value
        html = render_to_string(self.template_name, {'rows': rows, 'total_value': total_value})
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="inventory_valuation.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response


class InventoryExcelView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/inventory_excel.html'

    def get(self, request, *args, **kwargs):
        products = Product.objects.select_related('category').filter(is_active=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventory Valuation'
        ws.append(['Product', 'Category', 'SKU', 'Quantity', 'Cost Price', 'Value'])
        total_value = 0
        for product in products:
            value = product.cost_price * product.quantity_in_stock
            ws.append([product.name, product.category.name, product.sku, product.quantity_in_stock, product.cost_price, value])
            total_value += value
        ws.append([])
        ws.append(['Total Value', total_value])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventory_valuation.xlsx"'
        wb.save(response)
        return response
